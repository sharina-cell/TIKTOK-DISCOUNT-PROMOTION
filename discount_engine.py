"""
Marketplace Discount Promotion Upload File Generator
=====================================================

Supports Shopee and TikTok Seller Center reports today; built to extend to
Lazada / Zalora the same way (add a PLATFORM spec, nothing else changes).

Pipeline (same for every platform):
1. Parse the marketplace's raw report -> normalized product rows + live price
   + the SKU used to match against the Masterfile (platforms differ on which
   column that is, and on whether there's a fallback column).
2. Parse Masterfile (user-selected tab/columns) -> RRP + N discount tiers.
3. Merge on SKU, compute RRP CHECK, per-tier discount %, remarks.
4. Split out RRP mismatches into their own tab.
5. Build one "TO UPLOAD <TIER>" sheet per tier, in that platform's official
   upload template layout (these layouts differ a lot between platforms).

All marketplace xlsx exports we've seen have a broken `activePane` attribute
in their view XML that openpyxl chokes on - patch_activepane() strips it
before loading.
"""
import re
import io
import zipfile
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------------

def patch_activepane(file_bytes: bytes) -> bytes:
    """Strip invalid activePane="" attributes that break openpyxl on marketplace exports."""
    buf_in = io.BytesIO(file_bytes)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, "r") as zin, zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml"):
                data = re.sub(rb'activePane="[^"]*"', b"", data)
            zout.writestr(item, data)
    return buf_out.getvalue()


def load_workbook_safe(file_bytes: bytes, data_only: bool = True):
    """Load a workbook, auto-patching the activePane defect if needed."""
    try:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=data_only)
    except ValueError as e:
        if "activePane" in str(e) or "could not read worksheets" in str(e):
            return openpyxl.load_workbook(io.BytesIO(patch_activepane(file_bytes)), data_only=data_only)
        raise


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col_letter_to_idx(letter: str) -> int:
    """'F' -> 6"""
    return column_index_from_string(letter.strip().upper())


def _to_number(val):
    """Best-effort numeric coercion. Handles '$179.00', '179', 179, None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _clean_sku(val):
    """Normalize SKU for matching: strings, strip trailing .0 from float-ified numeric SKUs."""
    if val is None:
        return None
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    s = str(val).strip()
    if s == "":
        return None
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


# ---------------------------------------------------------------------------
# Platform specs
# ---------------------------------------------------------------------------
# "columns" maps a logical field name -> (header text to look for, fallback
# column letter if the header text isn't found on the detected header row).
# "match_key" is the logical field used to look products up in the Masterfile.
# "match_key_fallback" is used when match_key is blank for a row (optional).
# "upload_template" maps each output column header -> a logical field name
# from "columns", the special token "__tier_price__" (the SRP for the tier
# being built), or None (left blank, e.g. optional purchase-limit columns).

PLATFORM_SPECS = {
    "Shopee": {
        "header_markers": {"product id", "sku", "price"},
        "columns": {
            "product_id": ("Product ID", "A"),
            "product_name": ("Product Name", "B"),
            "variation_id": ("Variation ID", "C"),
            "variation_name": ("Variation Name", "D"),
            "parent_sku": ("Parent SKU", "E"),
            "sku": ("SKU", "F"),
            "price": ("Price", "G"),
            "gtin": ("GTIN", "H"),
            "stock": ("Stock", "I"),
        },
        "match_key": "sku",
        "match_key_fallback": "parent_sku",
        "display_order": ["product_id", "product_name", "variation_id", "variation_name",
                           "parent_sku", "sku", "price", "gtin", "stock"],
        "upload_template": {
            "Product ID": "product_id",
            "Product\xa0Name(Optional)": "product_name",
            "Parent\xa0SKU.\xa0Ref.\xa0No.(Optional)": "parent_sku",
            "Variation ID": "variation_id",
            "Variation\xa0name(Optional)": "variation_name",
            "SKU\xa0Ref.\xa0No.(Optional)": "sku",
            "Original price (Optional)": "price",
            "Discount price": "__tier_price__",
            "Purchase Limit (Optional)": None,
        },
    },
    "TikTok": {
        "header_markers": {"product id", "sku id", "seller sku"},
        "columns": {
            "product_id": ("Product ID", "A"),
            "category": ("Category", "B"),
            "product_name": ("Product name", "C"),
            "sku_id": ("SKU ID", "D"),
            "variation_option": ("Variation Option", "E"),
            "price": ("Retail Price (Local Currency)", "F"),
            "seller_sku": ("Seller SKU", "G"),
        },
        "match_key": "seller_sku",
        "match_key_fallback": None,
        "display_order": ["product_id", "category", "product_name", "sku_id",
                           "variation_option", "price", "seller_sku"],
        "upload_template": {
            "Product_id (required)": "product_id",
            "SKU_id (required)": "sku_id",
            "Deal Price (required)": "__tier_price__",
            "Total Purchase Limit (optional)": None,
            "Buyer purchase limit\uff08optional\uff09": None,
        },
    },
}


def list_platforms():
    return list(PLATFORM_SPECS.keys())


# ---------------------------------------------------------------------------
# Step 1: Marketplace report parsing (generic across platforms)
# ---------------------------------------------------------------------------

def find_header_row(ws, markers: set) -> int:
    """Marketplace exports often have junk rows above the real header. Detect it dynamically."""
    for r in range(1, min(ws.max_row, 20) + 1):
        row_vals = {
            str(ws.cell(r, c).value).strip().lower()
            for c in range(1, min(ws.max_column, 30) + 1)
            if ws.cell(r, c).value is not None
        }
        if markers.issubset(row_vals):
            return r
    raise ValueError(f"Could not find the report header row (expected columns like {sorted(markers)}).")


def parse_report(file_bytes: bytes, platform: str) -> pd.DataFrame:
    """
    Returns a DataFrame with one column per logical field defined in the
    platform's spec, plus a computed "SKU_MATCH" column (match_key, falling
    back to match_key_fallback when blank, per that platform's rule).
    """
    spec = PLATFORM_SPECS[platform]
    wb = load_workbook_safe(file_bytes)
    ws = wb.worksheets[0]
    header_row = find_header_row(ws, spec["header_markers"])

    headers = {str(ws.cell(header_row, c).value).strip(): c for c in range(1, ws.max_column + 1)
               if ws.cell(header_row, c).value is not None}

    field_cols = {}
    for field, (header_text, default_letter) in spec["columns"].items():
        field_cols[field] = headers.get(header_text, col_letter_to_idx(default_letter))

    c_product_id = field_cols.get("product_id")
    c_price = field_cols.get("price")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        pid = ws.cell(r, c_product_id).value if c_product_id else None
        if pid is None or str(pid).strip() == "":
            continue
        price_raw = ws.cell(r, c_price).value if c_price else None
        if isinstance(price_raw, str) and price_raw.strip().lower() in ("mandatory", ""):
            continue

        rec = {}
        for field, cidx in field_cols.items():
            val = ws.cell(r, cidx).value
            rec[field] = _to_number(val) if field == "price" else val
        rows.append(rec)

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    match_key = spec["match_key"]
    fallback_key = spec.get("match_key_fallback")
    if fallback_key:
        df["SKU_MATCH"] = df.apply(
            lambda row: _clean_sku(row.get(match_key)) or _clean_sku(row.get(fallback_key)), axis=1
        )
    else:
        df["SKU_MATCH"] = df[match_key].apply(_clean_sku)

    return df


# ---------------------------------------------------------------------------
# Step 2: Masterfile (identical across platforms)
# ---------------------------------------------------------------------------

def get_master_sheet_names(file_bytes: bytes):
    wb = load_workbook_safe(file_bytes)
    return wb.sheetnames


def get_master_headers(file_bytes: bytes, sheet_name: str, header_row: int = 1):
    """Returns list of (column_letter, header_text) for a quick picker in the UI."""
    wb = load_workbook_safe(file_bytes)
    ws = wb[sheet_name]
    out = []
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        val = ws.cell(header_row, c).value
        out.append((letter, val))
    return out


def parse_master_file(file_bytes: bytes, sheet_name: str, sku_col: str, rrp_col: str,
                       tiers: dict, header_row: int = 1) -> pd.DataFrame:
    """
    tiers: {tier_name: column_letter}, e.g. {"BAU": "F", "DDAY": "H", "SWFS": "I"}
    Returns DataFrame: SKU, RRP, <tier1>, <tier2>, ...
    """
    wb = load_workbook_safe(file_bytes)
    ws = wb[sheet_name]

    c_sku = col_letter_to_idx(sku_col)
    c_rrp = col_letter_to_idx(rrp_col)
    tier_cols = {name: col_letter_to_idx(letter) for name, letter in tiers.items()}

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        sku_val = ws.cell(r, c_sku).value
        if sku_val is None or str(sku_val).strip() == "":
            continue
        rec = {
            "SKU": _clean_sku(sku_val),
            "RRP": _to_number(ws.cell(r, c_rrp).value),
        }
        for name, cidx in tier_cols.items():
            rec[name] = _to_number(ws.cell(r, cidx).value)
        rows.append(rec)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Step 3-6: Merge + compute working sheet, mismatch tab, remarks
# ---------------------------------------------------------------------------

def build_working_sheet(report_df: pd.DataFrame, master_df: pd.DataFrame, tier_names: list,
                         platform: str) -> pd.DataFrame:
    """
    Left-join the report with master on SKU_MATCH. Computes:
      - RRP CHECK: master RRP == live price (both must exist)
      - <tier> DISCOUNT: 1 - tier_price / live price
      - remarks: "no discount remarks" if all available tier discounts are 0
    """
    spec = PLATFORM_SPECS[platform]
    df = report_df.merge(master_df, left_on="SKU_MATCH", right_on="SKU", how="left", suffixes=("", "_master"))

    df["RRP CHECK"] = df.apply(
        lambda row: (round(float(row["RRP"]), 2) == round(float(row["price"]), 2))
        if pd.notna(row.get("RRP")) and pd.notna(row.get("price")) else pd.NA,
        axis=1,
    )

    for tier in tier_names:
        disc_col = f"{tier} DISCOUNT"

        def calc(row, tier=tier):
            price = row.get("price")
            srp = row.get(tier)
            if pd.isna(price) or pd.isna(srp) or price in (None, 0):
                return pd.NA
            return 1 - (float(srp) / float(price))

        df[disc_col] = df.apply(calc, axis=1)

    def remark(row):
        vals = [row.get(f"{t} DISCOUNT") for t in tier_names]
        vals = [v for v in vals if pd.notna(v)]
        if vals and all(v == 0 for v in vals):
            return "no discount remarks"
        return None

    df["remarks"] = df.apply(remark, axis=1)

    ordered_cols = (
        list(spec["display_order"])
        + tier_names
        + ["RRP", "RRP CHECK"]
        + [f"{t} DISCOUNT" for t in tier_names]
        + ["remarks"]
    )
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    return df[ordered_cols]


def build_mismatch_sheet(working_df: pd.DataFrame) -> pd.DataFrame:
    """Rows where RRP CHECK is explicitly False (both RRP and live price present but differ)."""
    mask = working_df["RRP CHECK"].apply(lambda v: v is False)
    return working_df[mask].copy()


# ---------------------------------------------------------------------------
# Step 7: Final marketplace upload sheets (per tier), using that platform's
# official template header
# ---------------------------------------------------------------------------

def build_upload_sheet(working_df: pd.DataFrame, tier: str, platform: str) -> pd.DataFrame:
    """
    Builds the marketplace upload-ready sheet for one discount tier, per that
    platform's upload_template mapping. Only includes rows that have a valid
    (non-NA) price for that tier.
    """
    spec = PLATFORM_SPECS[platform]
    df = working_df[pd.notna(working_df[tier])].copy()
    out = pd.DataFrame()
    for out_col, source in spec["upload_template"].items():
        if source is None:
            out[out_col] = None
        elif source == "__tier_price__":
            out[out_col] = df[tier]
        else:
            out[out_col] = df[source]
    return out


# ---------------------------------------------------------------------------
# Step 8: Write everything to a single output workbook
# ---------------------------------------------------------------------------

def write_output_workbook(working_df: pd.DataFrame, mismatch_df: pd.DataFrame,
                           tier_names: list, working_sheet_df_for_upload: pd.DataFrame,
                           platform: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        working_df.to_excel(writer, sheet_name="Working Sheet", index=False)
        mismatch_df.to_excel(writer, sheet_name="RRP MISMATCH", index=False)
        for tier in tier_names:
            upload_df = build_upload_sheet(working_sheet_df_for_upload, tier, platform)
            safe_name = f"TO UPLOAD {tier}"[:31]
            upload_df.to_excel(writer, sheet_name=safe_name, index=False)
    buf.seek(0)
    return buf.getvalue()
